"""ChronoText benchmark scoring summary.

Aggregates rule-based judging results from ``judge_results/<model>/results.jsonl``
into a multi-sheet Excel workbook with per-model x per-task / per-font-type breakdowns.

Usage:
    python summarize.py                              # default: scan judge_results/
    python summarize.py --input_dir judge_results
    python summarize.py --output results_analysis.xlsx --num_workers 64
"""

from __future__ import annotations

import argparse
import json
import os
import sys
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from tqdm import tqdm

REPO_ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(REPO_ROOT.parent))

DEFAULT_INPUT_DIR = REPO_ROOT / "judge_results"
DEFAULT_OUTPUT_FILE = REPO_ROOT / "judge_results" / "results_analysis.xlsx"

# 数值格式：x100 保留 1 位（不加 %）
SCORE_SCALE = 100
SCORE_DECIMALS = 1
AVG_HEADER = "Average"

# 任务展示名：jsonl 中的 task key 用中文，输出表头统一映射为英文
TASK_DISPLAY = {
    "字体分类": "Classification",
    "字符提取": "Parsing",
    "字符检测_Detection": "Detection",
    "字符检测_Spotting": "Spotting",
    "单字识别": "Recognition",
}
DETECTION_TASK = "字符检测_Detection"
SPOTTING_TASK = "字符检测_Spotting"
RECOGNITION_TASK = "单字识别"

# 字体顺序：学术顺序，不要按字典序
FONT_TYPE_PRIORITY = ["甲骨文", "金文", "篆书", "隶书", "楷书", "行书", "草书"]
ANCIENT_FONTS = {"甲骨文", "金文", "篆书"}

ANCIENT_TASKS = ["字体分类", "字符提取", DETECTION_TASK, SPOTTING_TASK, RECOGNITION_TASK]
MODERN_TASKS = ["字体分类", "字符提取"]
DISPLAY_TASKS = [SPOTTING_TASK, RECOGNITION_TASK, "字符提取", "字体分类"]
_DISPLAY_ORDER = {t: i for i, t in enumerate(DISPLAY_TASKS)}


def display(t: str) -> str:
    return TASK_DISPLAY.get(t, t)


def filter_display(tasks: list[str]) -> list[str]:
    inter = [t for t in tasks if t in _DISPLAY_ORDER]
    inter.sort(key=lambda x: _DISPLAY_ORDER[x])
    return inter


def fmt(v) -> float | str:
    if v is None or v == "":
        return ""
    try:
        return round(float(v) * SCORE_SCALE, SCORE_DECIMALS)
    except (TypeError, ValueError):
        return ""


def parse_judge_file(file_path: str) -> tuple[dict[str, list[float]], dict[str, dict[str, list[float]]]]:
    """单次遍历同时返回整体分数 + 按字体分组的分数。"""
    task_scores: dict[str, list[float]] = defaultdict(list)
    type_task_scores: dict[str, dict[str, list[float]]] = defaultdict(lambda: defaultdict(list))
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    data = json.loads(line)
                except json.JSONDecodeError:
                    continue
                jr = data.get("judge_results") or {}
                if not jr:
                    continue
                font_type = str(data.get("font_type", "") or "").strip() or "未知"
                for task_name, task_result in jr.items():
                    # 字符检测拆成 Detection / Spotting 两个虚拟任务
                    if task_name == "字符检测":
                        inner = task_result.get("score", task_result) if isinstance(task_result, dict) else None
                        det = inner.get("detection_f1") if isinstance(inner, dict) else None
                        spot = inner.get("spotting_f1") if isinstance(inner, dict) else None
                        if det is not None:
                            task_scores[DETECTION_TASK].append(det)
                            type_task_scores[font_type][DETECTION_TASK].append(det)
                        if spot is not None:
                            task_scores[SPOTTING_TASK].append(spot)
                            type_task_scores[font_type][SPOTTING_TASK].append(spot)
                        continue
                    score = task_result.get("score", 0.0) if isinstance(task_result, dict) else 0.0
                    if isinstance(score, dict) and "score" in score:
                        score = score["score"]
                    task_scores[task_name].append(score)
                    type_task_scores[font_type][task_name].append(score)
    except Exception as e:
        print(f"读取文件 {file_path} 时出错: {e}")
    return task_scores, type_task_scores


def calc_avg(task_scores: dict[str, list]) -> dict[str, float | None]:
    out: dict[str, float | None] = {}
    for k, scores in task_scores.items():
        cleaned = []
        for s in scores:
            if isinstance(s, dict) and "score" in s:
                s = s["score"]
            if s is None:
                continue
            try:
                cleaned.append(float(s))
            except (TypeError, ValueError):
                continue
        out[k] = sum(cleaned) / len(cleaned) if cleaned else None
    return out


def font_allowed_tasks(font_type: str, available_tasks: list[str]) -> list[str]:
    allowed = ANCIENT_TASKS if font_type in ANCIENT_FONTS else MODERN_TASKS
    inter = [t for t in available_tasks if t in allowed]
    return filter_display(inter)


def get_group_tasks(group: str) -> list[str]:
    base = ANCIENT_TASKS if group == "ancient" else MODERN_TASKS if group == "modern" else []
    return filter_display(base)


def analyze(input_dir: str, output_file: str, num_workers: int) -> None:
    print("=" * 72)
    print("ChronoText Summarize")
    print("=" * 72)
    print(f"input_dir   : {input_dir}")
    print(f"output_file : {output_file}")
    print(f"num_workers : {num_workers}\n")

    if not os.path.isdir(input_dir):
        raise SystemExit(f"输入目录不存在: {input_dir}")

    model_tags = sorted(d for d in os.listdir(input_dir) if os.path.isdir(os.path.join(input_dir, d)))
    print(f"找到 {len(model_tags)} 个模型: {model_tags}\n")

    tasks: list[tuple[str, str]] = []
    for tag in model_tags:
        f = os.path.join(input_dir, tag, "results.jsonl")
        if os.path.isfile(f):
            tasks.append((tag, f))
        else:
            print(f"  跳过 {tag}：找不到 {f}")

    per_model_overall: dict[str, dict[str, float | None]] = {}
    per_model_by_font: dict[str, dict[str, dict[str, float | None]]] = {}
    per_model_count: dict[str, dict[str, int]] = {}
    all_tasks: set[str] = set()
    all_fonts: set[str] = set()

    def _worker(item):
        tag, fpath = item
        ts, tts = parse_judge_file(fpath)
        return tag, calc_avg(ts), tts

    workers = max(1, min(num_workers, len(tasks))) if tasks else 1
    with ThreadPoolExecutor(max_workers=workers) as ex:
        futs = [ex.submit(_worker, t) for t in tasks]
        for fut in tqdm(as_completed(futs), total=len(futs), desc="parse jsonl"):
            tag, overall, tts = fut.result()
            per_model_overall[tag] = overall
            all_tasks.update(overall.keys())
            font_avg: dict[str, dict[str, float | None]] = {}
            cnt: dict[str, int] = {}
            for ft, tmap in tts.items():
                font_avg[ft] = {tn: (sum(v) / len(v) if v else None) for tn, v in tmap.items()}
                cnt[ft] = max((len(v) for v in tmap.values()), default=0)
            per_model_by_font[tag] = font_avg
            per_model_count[tag] = cnt
            all_fonts.update(tts.keys())

    sorted_models = sorted(per_model_overall.keys())  # 按字典序展示
    sorted_fonts = [f for f in FONT_TYPE_PRIORITY if f in all_fonts] + sorted(all_fonts - set(FONT_TYPE_PRIORITY))

    ancient_tasks = get_group_tasks("ancient")
    modern_tasks = get_group_tasks("modern")

    # ==================== Sheet 1: 评分分析（古代 / 近代汇总） ====================
    rows = []
    for tag in sorted_models:
        font_avg = per_model_by_font.get(tag, {})

        row: dict[str, object] = {"模型名称": tag}

        # 古代区块：只取古代字体下的 per-font 均分，再对字体求平均
        anc_scores: dict[str, list[float]] = defaultdict(list)
        for ft in sorted_fonts:
            if ft not in ANCIENT_FONTS:
                continue
            for t, v in font_avg.get(ft, {}).items():
                if t in ancient_tasks and v is not None:
                    anc_scores[t].append(v)
        anc_per = {t: (sum(v) / len(v) if v else None) for t, v in anc_scores.items()}
        anc_valid = [anc_per.get(t) for t in ancient_tasks if anc_per.get(t) is not None]
        row["平均分古代_avg"] = fmt(sum(anc_valid) / len(anc_valid) if anc_valid else None)
        for t in ancient_tasks:
            row[f"平均分古代_{t}"] = fmt(anc_per.get(t))

        # 近代区块：只取近代字体下的 per-font 均分，再对字体求平均
        mod_scores: dict[str, list[float]] = defaultdict(list)
        for ft in sorted_fonts:
            if ft in ANCIENT_FONTS:
                continue
            for t, v in font_avg.get(ft, {}).items():
                if t in modern_tasks and v is not None:
                    mod_scores[t].append(v)
        mod_per = {t: (sum(v) / len(v) if v else None) for t, v in mod_scores.items()}
        mod_valid = [mod_per.get(t) for t in modern_tasks if mod_per.get(t) is not None]
        row["平均分近代_avg"] = fmt(sum(mod_valid) / len(mod_valid) if mod_valid else None)
        for t in modern_tasks:
            row[f"平均分近代_{t}"] = fmt(mod_per.get(t))

        rows.append(row)

    df = pd.DataFrame(rows)
    header1 = ["模型名称"] + ["平均分_古代"] * (len(ancient_tasks) + 1) + ["平均分_近代"] * (len(modern_tasks) + 1)
    header2 = (
        ["模型名称", AVG_HEADER]
        + [display(t) for t in ancient_tasks]
        + [AVG_HEADER]
        + [display(t) for t in modern_tasks]
    )

    column_order = (
        ["模型名称", "平均分古代_avg"]
        + [f"平均分古代_{t}" for t in ancient_tasks]
        + ["平均分近代_avg"]
        + [f"平均分近代_{t}" for t in modern_tasks]
    )
    for col in column_order:
        if col not in df.columns:
            df[col] = ""
    df = df[column_order]

    os.makedirs(os.path.dirname(output_file) or ".", exist_ok=True)
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="评分分析", index=False, startrow=2, header=False)
        ws = writer.sheets["评分分析"]
        for ci, v in enumerate(header1, start=1):
            ws.cell(row=1, column=ci, value=v)
        for ci, v in enumerate(header2, start=1):
            ws.cell(row=2, column=ci, value=v)
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
        ci = 2
        ws.merge_cells(start_row=1, start_column=ci, end_row=1, end_column=ci + len(ancient_tasks))
        ci += len(ancient_tasks) + 1
        ws.merge_cells(start_row=1, start_column=ci, end_row=1, end_column=ci + len(modern_tasks))

        head_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        head_font = Font(bold=True)
        center = Alignment(horizontal="center", vertical="center")
        for r in (1, 2):
            for c in range(1, len(header2) + 1):
                cell = ws.cell(row=r, column=c)
                cell.fill = head_fill
                cell.font = head_font
                cell.alignment = center
        ws.column_dimensions["A"].width = 30
        for c in range(2, len(header2) + 1):
            ws.column_dimensions[get_column_letter(c)].width = 12

    # ==================== Sheet 2: 按字体分析 ====================
    if sorted_fonts:
        # 顶部 Average 区域只展示出现过的、属于古代∪近代任一组、且在 DISPLAY_TASKS 内的任务
        all_visible = filter_display([t for t in all_tasks if t in (set(ANCIENT_TASKS) | set(MODERN_TASKS))])

        type_rows = []
        for tag in sorted_models:
            font_avg = per_model_by_font.get(tag, {})
            cnt = per_model_count.get(tag, {})
            row: dict[str, object] = {"模型名称": tag}

            scores_by_task: dict[str, list[float]] = defaultdict(list)
            all_for_avg: list[float] = []
            for ft, tmap in font_avg.items():
                allowed = set(font_allowed_tasks(ft, sorted(all_tasks)))
                for t, v in tmap.items():
                    if t.startswith("_") or t not in allowed or v is None:
                        continue
                    scores_by_task[t].append(v)
                    all_for_avg.append(v)
            row["平均分_avg"] = fmt(sum(all_for_avg) / len(all_for_avg) if all_for_avg else None)
            for t in all_visible:
                vs = scores_by_task.get(t, [])
                row[f"平均分_{t}"] = fmt(sum(vs) / len(vs) if vs else None)

            for ft in sorted_fonts:
                tmap = font_avg.get(ft, {})
                ft_tasks = font_allowed_tasks(ft, sorted(all_tasks))
                allowed_set = set(ft_tasks)
                valid = [v for k, v in tmap.items() if k in allowed_set and v is not None]
                row[f"{ft}_avg"] = fmt(sum(valid) / len(valid) if valid else None)
                for t in ft_tasks:
                    v = tmap.get(t)
                    row[f"{ft}_{t}"] = fmt(v) if v is not None else ""
            type_rows.append(row)

        df_t = pd.DataFrame(type_rows)
        type_header1 = ["模型名称"] + ["平均分"] * (len(all_visible) + 1)
        for ft in sorted_fonts:
            type_header1.extend([ft] * (len(font_allowed_tasks(ft, sorted(all_tasks))) + 1))
        type_header2 = ["模型名称", AVG_HEADER] + [display(t) for t in all_visible]
        for ft in sorted_fonts:
            ft_tasks = font_allowed_tasks(ft, sorted(all_tasks))
            type_header2.append(AVG_HEADER)
            type_header2.extend(display(t) for t in ft_tasks)

        type_columns = ["模型名称", "平均分_avg"] + [f"平均分_{t}" for t in all_visible]
        for ft in sorted_fonts:
            ft_tasks = font_allowed_tasks(ft, sorted(all_tasks))
            type_columns.append(f"{ft}_avg")
            type_columns.extend(f"{ft}_{t}" for t in ft_tasks)
        for col in type_columns:
            if col not in df_t.columns:
                df_t[col] = ""
        df_t = df_t[type_columns]

        with pd.ExcelWriter(output_file, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            df_t.to_excel(writer, sheet_name="按字体分析", index=False, startrow=2, header=False)
            ws_t = writer.sheets["按字体分析"]
            for ci, v in enumerate(type_header1, start=1):
                ws_t.cell(row=1, column=ci, value=v)
            for ci, v in enumerate(type_header2, start=1):
                ws_t.cell(row=2, column=ci, value=v)
            ws_t.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
            ci = 2
            ws_t.merge_cells(start_row=1, start_column=ci, end_row=1, end_column=ci + len(all_visible))
            ci += len(all_visible) + 1
            for ft in sorted_fonts:
                span = len(font_allowed_tasks(ft, sorted(all_tasks))) + 1
                ws_t.merge_cells(start_row=1, start_column=ci, end_row=1, end_column=ci + span - 1)
                ci += span
            head_fill_t = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
            for r in (1, 2):
                for c in range(1, len(type_header2) + 1):
                    cell = ws_t.cell(row=r, column=c)
                    cell.fill = head_fill_t
                    cell.font = head_font
                    cell.alignment = center
            ws_t.column_dimensions["A"].width = 30
            for c in range(2, len(type_header2) + 1):
                ws_t.column_dimensions[get_column_letter(c)].width = 12

    print(f"\n✅ 已写入 {output_file}")


def main() -> None:
    p = argparse.ArgumentParser(description="ChronoText scoring summary")
    p.add_argument("--input_dir", type=str, default=str(DEFAULT_INPUT_DIR))
    p.add_argument("--output", type=str, default=str(DEFAULT_OUTPUT_FILE))
    p.add_argument("--num_workers", type=int, default=32)
    args = p.parse_args()
    analyze(args.input_dir, args.output, args.num_workers)


if __name__ == "__main__":
    main()
